from pathlib import Path
from openpyxl import load_workbook
import re, hashlib

repo = Path('/Users/user/Frontend-Development-')
wb_path = Path('/Users/user/Downloads/Frontend_Development_Practice_Bank.xlsx')
wb = load_workbook(wb_path)
ws = wb['Sheet1']

img_root = repo / 'solution-images-contextual'
img_root.mkdir(parents=True, exist_ok=True)

owner='Nithya-sri-14'
repo_name='Frontend-Development-'
branch='main'

def slugify(t):
    return re.sub(r'[^a-z0-9]+','-',t.lower()).strip('-')

def pick_palette(topic):
    return {
        'HTML': ('#0f766e','#f0fdfa'),'Semantic HTML': ('#0369a1','#f0f9ff'),'Forms': ('#1d4ed8','#eff6ff'),'Tables': ('#334155','#f8fafc'),
        'CSS': ('#7c3aed','#f5f3ff'),'Box Model': ('#b45309','#fff7ed'),'Flexbox': ('#0e7490','#ecfeff'),'Grid': ('#be185d','#fdf2f8'),
        'Media Queries': ('#166534','#f0fdf4'),'Responsive Design': ('#9a3412','#fff7ed')
    }.get(topic,('#1f2937','#f8fafc'))

def layout(seed):
    h = int(hashlib.md5(seed.encode()).hexdigest()[:8], 16)
    p = h % 4
    if p == 0:
        return [(70,170,1140,90),(70,275,560,170),(650,275,560,170),(70,460,760,180),(850,460,360,180)]
    if p == 1:
        return [(70,170,1140,100),(70,285,360,355),(450,285,760,110),(450,410,370,230),(840,410,370,230)]
    if p == 2:
        return [(70,170,1140,110),(70,295,1140,120),(70,430,360,210),(450,430,360,210),(830,430,380,210)]
    return [(70,170,760,120),(850,170,360,120),(70,305,360,335),(450,305,360,335),(830,305,380,335)]

def sections_for(title, topic):
    t = title.lower()
    # focused contextual templates
    if 'restaurant' in t:
        return [
            ('Welcome Banner', 'Restaurant name, short tagline, and hero image area'),
            ('Menu Highlights', 'Signature dishes with price snippets and cuisine tags'),
            ('Chef Special', 'Featured meal card with ingredients and serving note'),
            ('Reservation', 'Date/time/guest call-to-action with booking button'),
            ('Contact & Hours', 'Address, open timings, and phone number section')
        ]
    if 'portfolio' in t:
        return [
            ('Top Navigation', 'Brand, links to About, Projects, and Contact'),
            ('Hero Introduction', 'Name, role, short summary, and primary action'),
            ('About Section', 'Skills, experience highlights, and profile note'),
            ('Projects Grid', 'Project cards with title, stack, and preview links'),
            ('Contact Block', 'Email, social links, and collaboration CTA')
        ]
    if 'form' in t or topic == 'Forms':
        return [
            ('Form Header', 'Purpose statement and instruction summary'),
            ('Personal Details', 'Name, email, phone, and required inputs'),
            ('Selection Inputs', 'Dropdown/radio/checkbox option groups'),
            ('Validation Area', 'Error/helper messages and field constraints'),
            ('Submission Actions', 'Submit/reset controls and consent note')
        ]
    if 'table' in t or topic == 'Tables':
        return [
            ('Table Caption', 'Dataset name and context summary'),
            ('Column Headers', 'Structured labels for each data dimension'),
            ('Data Rows', 'Representative entries aligned in readable columns'),
            ('Summary Row', 'Total/average/key metrics with emphasis'),
            ('Legend Notes', 'Status meaning and interpretation guide')
        ]
    if 'dashboard' in t:
        return [
            ('Header Controls', 'Date filter, search, and quick actions'),
            ('KPI Cards', 'Primary metrics with trend indicators'),
            ('Chart Area', 'Visualization block for weekly/monthly trend'),
            ('Detail Panel', 'Breakdown widgets and category stats'),
            ('Recent Activity', 'Latest updates and status list')
        ]
    if topic == 'Semantic HTML':
        return [
            ('Page Header', 'Title, author/date meta, and intro summary'),
            ('Main Article', 'Primary narrative content with clear hierarchy'),
            ('Supporting Aside', 'Contextual references and quick facts'),
            ('Figure + Caption', 'Illustration area linked to article context'),
            ('Footer Meta', 'Source links and publication details')
        ]
    if topic == 'Grid':
        return [
            ('Top Region', 'Main heading and control strip'),
            ('Grid Area A', 'Primary content panel'),
            ('Grid Area B', 'Secondary content panel'),
            ('Grid Area C', 'Metrics/details panel'),
            ('Bottom Region', 'Summary or footer actions')
        ]
    if topic == 'Flexbox':
        return [
            ('Flex Header', 'Horizontal nav and utility controls'),
            ('Aligned Row', 'Items distributed with spacing rules'),
            ('Action Group', 'Buttons aligned with consistent gaps'),
            ('Wrapped Items', 'Responsive wrap behavior preview'),
            ('Footer Row', 'Final aligned links/actions')
        ]
    # default contextual from keywords
    words = [w.capitalize() for w in re.split(r'[^a-z0-9]+', t) if w][:8]
    core = ' '.join(words[:3]) if words else 'Page'
    return [
        (f'{core} Header', 'Primary heading area and top navigation'),
        (f'{core} Overview', 'Main information block matching task intent'),
        (f'{core} Feature', 'Detailed section with supporting content'),
        (f'{core} Details', 'Secondary content with structured layout'),
        (f'{core} Footer', 'Contact/actions/closing summary area')
    ]

for r in range(2, ws.max_row+1):
    level = ws.cell(r,2).value
    topic = ws.cell(r,3).value
    title = ws.cell(r,4).value

    accent,bg = pick_palette(topic)
    rects = layout(f'{level}|{topic}|{title}')
    sections = sections_for(title, topic)
    fills = ['#dbeafe','#e2e8f0','#fef3c7','#dcfce7','#ede9fe']

    parts = [
        f"<svg xmlns='http://www.w3.org/2000/svg' width='1280' height='720' viewBox='0 0 1280 720'>",
        f"<rect width='1280' height='720' fill='{bg}'/>",
        f"<rect x='28' y='28' width='1224' height='664' rx='20' fill='white' stroke='{accent}' stroke-width='4'/>",
        f"<text x='60' y='80' font-family='Arial, sans-serif' font-size='28' font-weight='800' fill='{accent}'>{title}</text>",
    ]

    for i,(x,y,w,h) in enumerate(rects):
        htxt, ptxt = sections[i]
        parts.append(f"<rect x='{x}' y='{y}' width='{w}' height='{h}' rx='12' fill='{fills[i]}'/>")
        parts.append(f"<text x='{x+14}' y='{y+34}' font-family='Arial, sans-serif' font-size='21' font-weight='700' fill='#111827'>{htxt}</text>")
        # wrap to two lines
        line1 = ptxt[:56]
        line2 = ptxt[56:112]
        parts.append(f"<text x='{x+14}' y='{y+62}' font-family='Arial, sans-serif' font-size='15' fill='#1f2937'>{line1}</text>")
        if line2:
            parts.append(f"<text x='{x+14}' y='{y+82}' font-family='Arial, sans-serif' font-size='15' fill='#1f2937'>{line2}</text>")

    parts.append('</svg>')

    out_dir = img_root / slugify(level) / slugify(topic)
    out_dir.mkdir(parents=True, exist_ok=True)
    fname = f"q{r-1:03d}-{slugify(title)[:100]}.svg"
    fpath = out_dir / fname
    fpath.write_text(''.join(parts), encoding='utf-8')

    rel = fpath.relative_to(repo).as_posix()
    link = f"https://raw.githubusercontent.com/{owner}/{repo_name}/{branch}/{rel}"
    ws.cell(r,9).value = link
    ws.cell(r,10).value = link

wb.save(wb_path)
print('Generated contextual output images and updated links in workbook')
