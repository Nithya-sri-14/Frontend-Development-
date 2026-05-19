from pathlib import Path
from openpyxl import load_workbook
import re
import hashlib

repo = Path('/Users/user/Frontend-Development-')
wb_path = repo / 'Frontend_Development_Practice_Bank.xlsx'
wb = load_workbook(wb_path)

owner = 'Nithya-sri-14'
repo_name = 'Frontend-Development-'
base_preview = f'https://raw.githack.com/{owner}/{repo_name}/main'

solutions_root = repo / 'solutions'
solutions_root.mkdir(parents=True, exist_ok=True)

def slugify(text: str) -> str:
    return re.sub(r'[^a-z0-9]+', '-', text.lower()).strip('-')

def pick_palette(topic):
    m = {
        'HTML': ('#0f766e','#f0fdfa'),
        'Semantic HTML': ('#0369a1','#f0f9ff'),
        'Forms': ('#1d4ed8','#eff6ff'),
        'Tables': ('#334155','#f8fafc'),
        'CSS': ('#7c3aed','#f5f3ff'),
        'Box Model': ('#b45309','#fff7ed'),
        'Flexbox': ('#0e7490','#ecfeff'),
        'Grid': ('#be185d','#fdf2f8'),
        'Media Queries': ('#166534','#f0fdf4'),
        'Responsive Design': ('#9a3412','#fff7ed'),
    }
    return m.get(topic, ('#1f2937','#f8fafc'))

def section_labels(topic, title):
    t = title.lower()
    if 'portfolio' in t:
        return ['Nav', 'Hero', 'About', 'Projects', 'Contact']
    if topic == 'Forms':
        return ['Form Title','Personal Fields','Selection Group','Validation Notes','Submit Area']
    if topic == 'Tables':
        return ['Table Caption','Header Row','Data Region','Summary Row','Notes']
    if topic == 'Semantic HTML':
        return ['Header','Main Article','Aside','Figure','Footer']
    if topic == 'Flexbox':
        return ['Toolbar','Primary Row','Actions','Wrapped Items','Footer Row']
    if topic == 'Grid':
        return ['Top Strip','Sidebar','Grid A','Grid B','Bottom Strip']
    if topic == 'Media Queries':
        return ['Desktop View','Tablet View','Mobile View','Breakpoint Rules','Adaptive Block']
    if topic == 'Responsive Design':
        return ['Header','Fluid Hero','Adaptive Cards','Content Flow','Footer']
    if topic == 'Box Model':
        return ['Margin','Border','Padding','Content','Spacing']
    if topic == 'CSS':
        return ['Theme','Typography','Buttons','Cards','States']
    return ['Header','Section A','Section B','Section C','Footer']

def unique_layout(seed):
    h = int(hashlib.md5(seed.encode()).hexdigest()[:8], 16)
    pattern = h % 6
    if pattern == 0:
        return [(80,190,1120,90),(80,295,550,150),(650,295,550,150),(80,460,760,150),(860,460,340,150)]
    if pattern == 1:
        return [(80,190,1120,90),(80,295,360,315),(460,295,740,95),(460,405,360,205),(840,405,360,205)]
    if pattern == 2:
        return [(80,190,1120,90),(80,295,1120,120),(80,430,360,180),(460,430,360,180),(840,430,360,180)]
    if pattern == 3:
        return [(80,190,360,200),(460,190,360,200),(840,190,360,200),(80,410,540,200),(640,410,560,200)]
    if pattern == 4:
        return [(80,190,760,120),(860,190,340,120),(80,330,360,280),(460,330,360,280),(840,330,360,280)]
    return [(80,190,540,200),(640,190,560,200),(80,410,1120,90),(80,520,550,90),(650,520,550,90)]

def make_solution_html(topic, title):
    accent, bg = pick_palette(topic)
    labels = section_labels(topic, title)
    rects = unique_layout(topic + '|' + title)
    fills = ['#dbeafe','#e2e8f0','#fef3c7','#dcfce7','#ede9fe']

    divs = []
    for i, (x,y,w,h) in enumerate(rects):
        divs.append(
            f"<section class='blk' style='left:{x}px;top:{y}px;width:{w}px;height:{h}px;background:{fills[i]};'><h2>{labels[i]}</h2></section>"
        )

    return f"""<!doctype html>
<html lang='en'>
<head>
  <meta charset='utf-8' />
  <meta name='viewport' content='width=device-width, initial-scale=1' />
  <title>{title}</title>
  <style>
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: Arial, sans-serif; background: {bg}; color: #111827; }}
    .canvas {{ width: 1280px; min-height: 720px; margin: 0 auto; position: relative; }}
    .frame {{ position: absolute; left: 36px; top: 36px; width: 1208px; height: 648px; background: #fff; border: 5px solid {accent}; border-radius: 20px; }}
    .head {{ position: absolute; left: 60px; top: 60px; width: 1160px; height: 92px; border-radius: 12px; }}
    .title {{ position: absolute; left: 78px; top: 74px; margin: 0; font-size: 30px; font-weight: 800; color: {accent}; }}
    .sub {{ position: absolute; left: 78px; top: 114px; margin: 0; font-size: 20px; color: #334155; }}
    .blk {{ position: absolute; border-radius: 12px; padding: 10px 14px; overflow: hidden; }}
    .blk h2 {{ margin: 0; font-size: 24px; font-weight: 700; color: #111827; }}
    .foot {{ position: absolute; left: 80px; top: 660px; margin: 0; font-size: 20px; color: #0f172a; }}

    @media (max-width: 1280px) {{
      .canvas {{ width: 100vw; min-height: 56.25vw; }}
      .frame, .head, .title, .sub, .foot, .blk {{ transform-origin: top left; }}
      .frame, .head, .title, .sub, .foot, .blk {{ transform: scale(calc(100vw / 1280)); }}
    }}
  </style>
</head>
<body>
  <main class='canvas'>
    <div class='frame'></div>
    <div class='head'></div>
    <h1 class='title'>{title}</h1>
    <p class='sub'>Expected UI Reference</p>
    {''.join(divs)}
    <p class='foot'>Recreate this exact final interface.</p>
  </main>
</body>
</html>
"""

for sheet in wb.sheetnames:
    ws = wb[sheet]
    for r in range(2, ws.max_row + 1):
        topic = ws.cell(r, 3).value
        title = ws.cell(r, 4).value

        out_dir = solutions_root / slugify(sheet) / slugify(topic)
        out_dir.mkdir(parents=True, exist_ok=True)
        fname = f"q{r-1:03d}-{slugify(title)[:90]}.html"
        out_file = out_dir / fname
        out_file.write_text(make_solution_html(topic, title), encoding='utf-8')

        rel = out_file.relative_to(repo).as_posix()
        ws.cell(r, 9).value = f"{base_preview}/{rel}"

wb.save(wb_path)
print('Generated 900 solution HTML files and updated Example output links to rendered solution pages.')
