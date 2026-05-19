from pathlib import Path
from openpyxl import load_workbook
import re, hashlib

repo = Path('/Users/user/Frontend-Development-')
xlsx = Path('/Users/user/Documents/Frontend_Development_Practice_Bank.xlsx')
wb = load_workbook(xlsx)
ws = wb['Sheet1']

img_root = repo / 'solution-images-contextual-clean'
img_root.mkdir(parents=True, exist_ok=True)
owner='Nithya-sri-14'; repo_name='Frontend-Development-'; branch='main'

def slugify(t):
    return re.sub(r'[^a-z0-9]+','-',t.lower()).strip('-')

def strip_suffix(title):
    return re.sub(r'\s-\s[^-]*Practice\s[BIH]\d{3}$', '', title).strip()

def pick_palette(topic):
    return {
        'HTML': ('#0f766e','#f0fdfa'),'Semantic HTML': ('#0369a1','#f0f9ff'),'Forms': ('#1d4ed8','#eff6ff'),'Tables': ('#334155','#f8fafc'),
        'CSS': ('#7c3aed','#f5f3ff'),'Box Model': ('#b45309','#fff7ed'),'Flexbox': ('#0e7490','#ecfeff'),'Grid': ('#be185d','#fdf2f8'),
        'Media Queries': ('#166534','#f0fdf4'),'Responsive Design': ('#9a3412','#fff7ed')
    }.get(topic,('#1f2937','#f8fafc'))

def layout(seed):
    h = int(hashlib.md5(seed.encode()).hexdigest()[:8], 16)
    p = h % 4
    if p == 0:
        return [(70,170,1140,90),(70,275,560,170),(650,275,560,170),(70,460,760,180),(850,460,360,180)]
    if p == 1:
        return [(70,170,1140,100),(70,285,360,355),(450,285,760,110),(450,410,370,230),(840,410,370,230)]
    if p == 2:
        return [(70,170,1140,110),(70,295,1140,120),(70,430,360,210),(450,430,360,210),(830,430,380,210)]
    return [(70,170,760,120),(850,170,360,120),(70,305,360,335),(450,305,360,335),(830,305,380,335)]

def sections_for(title, topic):
    t = title.lower()
    if 'restaurant' in t:
        return [('Welcome Banner','Restaurant name, tagline, and hero image area'),('Menu Highlights','Signature dishes and category highlights'),('Chef Special','Featured plate with details'),('Reservation','Booking CTA with date/time block'),('Contact & Hours','Address, timings, and contact information')]
    if 'portfolio' in t:
        return [('Top Navigation','Brand and section links'),('Hero Introduction','Name, role, and short bio'),('About Section','Skills and experience highlights'),('Projects Grid','Project cards with brief summaries'),('Contact Block','Email/social links and CTA')]
    if topic == 'Forms':
        return [('Form Header','Purpose statement and guidance'),('Personal Details','Labelled inputs for user details'),('Selection Inputs','Dropdown/radio/checkbox groups'),('Validation Area','Error/help messaging patterns'),('Submission Actions','Submit/reset and consent controls')]
    if topic == 'Tables':
        return [('Table Caption','Dataset title and context'),('Column Headers','Structured field labels'),('Data Rows','Representative row entries'),('Summary Row','Totals/aggregates section'),('Legend Notes','Interpretation notes')]
    words=[w.capitalize() for w in re.split(r'[^a-z0-9]+',t) if w]
    core=' '.join(words[:3]) if words else 'Page'
    return [(f'{core} Header','Top information area'),(f'{core} Overview','Primary content section'),(f'{core} Feature','Detailed sub-section'),(f'{core} Details','Supporting information panel'),(f'{core} Footer','Closing actions and links')]

for r in range(2, ws.max_row+1):
    level = ws.cell(r,2).value
    topic = ws.cell(r,3).value
    title = ws.cell(r,4).value or ''
    clean = strip_suffix(title)
    ws.cell(r,4).value = clean

    accent,bg = pick_palette(topic)
    rects = layout(f'{level}|{topic}|{clean}')
    sections = sections_for(clean, topic)
    fills = ['#dbeafe','#e2e8f0','#fef3c7','#dcfce7','#ede9fe']

    parts = [
        "<svg xmlns='http://www.w3.org/2000/svg' width='1280' height='720' viewBox='0 0 1280 720'>",
        f"<rect width='1280' height='720' fill='{bg}'/>",
        f"<rect x='28' y='28' width='1224' height='664' rx='20' fill='white' stroke='{accent}' stroke-width='4'/>",
        f"<text x='60' y='80' font-family='Arial, sans-serif' font-size='28' font-weight='800' fill='{accent}'>{clean}</text>",
    ]
    for i,(x,y,w,h) in enumerate(rects):
        htxt,ptxt=sections[i]
        parts.append(f"<rect x='{x}' y='{y}' width='{w}' height='{h}' rx='12' fill='{fills[i]}'/>")
        parts.append(f"<text x='{x+14}' y='{y+34}' font-family='Arial, sans-serif' font-size='21' font-weight='700' fill='#111827'>{htxt}</text>")
        parts.append(f"<text x='{x+14}' y='{y+60}' font-family='Arial, sans-serif' font-size='15' fill='#1f2937'>{ptxt[:70]}</text>")

    parts.append('</svg>')
    out_dir = img_root / slugify(level) / slugify(topic)
    out_dir.mkdir(parents=True, exist_ok=True)
    fname = f"q{r-1:03d}-{slugify(clean)[:100]}.svg"
    fp = out_dir / fname
    fp.write_text(''.join(parts), encoding='utf-8')

    rel = fp.relative_to(repo).as_posix()
    link = f"https://raw.githubusercontent.com/{owner}/{repo_name}/{branch}/{rel}"
    ws.cell(r,9).value = link
    ws.cell(r,10).value = link

wb.save(xlsx)
print('Updated titles and contextual output image links:', xlsx)
