from pathlib import Path
from openpyxl import load_workbook
import re, hashlib

repo = Path('/Users/user/Frontend-Development-')
xlsx = Path('/Users/user/Documents/Frontend_Development_Practice_Bank.xlsx')
wb = load_workbook(xlsx)
ws = wb['Sheet1']

img_root = repo / 'solution-images-problem-matched'
img_root.mkdir(parents=True, exist_ok=True)

owner='Nithya-sri-14'; repo_name='Frontend-Development-'; branch='main'

def slugify(t):
    return re.sub(r'[^a-z0-9]+','-',t.lower()).strip('-')

def palette(topic):
    return {
        'HTML': ('#0f766e','#f0fdfa'),'Semantic HTML': ('#0369a1','#f0f9ff'),'Forms': ('#1d4ed8','#eff6ff'),'Tables': ('#334155','#f8fafc'),
        'CSS': ('#7c3aed','#f5f3ff'),'Box Model': ('#b45309','#fff7ed'),'Flexbox': ('#0e7490','#ecfeff'),'Grid': ('#be185d','#fdf2f8'),
        'Media Queries': ('#166534','#f0fdf4'),'Responsive Design': ('#9a3412','#fff7ed')
    }.get(topic,('#1f2937','#f8fafc'))

def layout(seed):
    h = int(hashlib.md5(seed.encode()).hexdigest()[:8], 16)
    p = h % 3
    if p == 0:
        return [(64,160,1152,92),(64,268,560,180),(656,268,560,180),(64,464,760,192),(848,464,368,192)]
    if p == 1:
        return [(64,160,1152,100),(64,276,368,380),(448,276,768,116),(448,408,376,248),(840,408,376,248)]
    return [(64,160,1152,108),(64,284,1152,128),(64,428,368,228),(448,428,368,228),(832,428,384,228)]

def lines_for_problem(topic, title):
    t = title.lower()
    if topic == 'Forms' or 'form' in t or 'application' in t:
        return [
            ('Form Header', ['Application Form', 'Complete all mandatory fields']),
            ('Personal Details', ['Full Name', 'Address', 'Email', 'Phone Number']),
            ('Additional Info', ['Date of Birth', 'City/State', 'Category/Role']),
            ('Validation Rules', ['Required fields (*)', 'Email format and phone length']),
            ('Submission Block', ['Submit Application', 'Reset Form'])
        ]
    if topic == 'Tables' or 'table' in t:
        return [
            ('Table Caption', ['Dataset: performance summary']),
            ('Headers', ['ID | Name | Category | Score | Status']),
            ('Rows', ['Sample row values aligned in columns']),
            ('Summary', ['Total, average, and status count']),
            ('Notes', ['Legend for status colors'])
        ]
    if 'restaurant' in t:
        return [
            ('Welcome Section', ['Restaurant Name', 'Tagline and hero image']),
            ('Menu Highlights', ['Starters', 'Main Course', 'Desserts']),
            ('Special Offer', ['Chef Special of the Day']),
            ('Reservation', ['Date', 'Time', 'Guests', 'Book Table']),
            ('Contact & Hours', ['Address', 'Phone', 'Open Hours'])
        ]
    if 'portfolio' in t:
        return [
            ('Navigation', ['Home | About | Projects | Contact']),
            ('Hero Intro', ['Name', 'Role', 'Short professional summary']),
            ('About', ['Skills', 'Experience', 'Education']),
            ('Projects', ['Project 1', 'Project 2', 'Project 3']),
            ('Contact', ['Email', 'LinkedIn', 'GitHub'])
        ]
    if topic == 'Semantic HTML':
        return [
            ('Header Landmark', ['Site title', 'Article metadata']),
            ('Main Article', ['Structured headings and paragraphs']),
            ('Aside', ['Related links and context notes']),
            ('Figure', ['Image region + figcaption']),
            ('Footer', ['Author and source details'])
        ]
    if topic == 'Grid':
        return [
            ('Top Grid Area', ['Heading and controls']),
            ('Area A', ['Primary content cards']),
            ('Area B', ['Secondary information']),
            ('Area C', ['Metrics / details panel']),
            ('Bottom Area', ['Summary / actions'])
        ]
    if topic == 'Flexbox':
        return [
            ('Flex Header', ['Logo', 'Nav links', 'Profile action']),
            ('Aligned Row', ['Items distributed with justify-content']),
            ('Action Group', ['Primary and secondary buttons']),
            ('Wrapped Block', ['Tags/chips wrap to next line']),
            ('Footer Row', ['Aligned links and status'])
        ]
    if topic == 'Media Queries':
        return [
            ('Desktop Layout', ['3-column arrangement']),
            ('Tablet Layout', ['2-column reflow']),
            ('Mobile Layout', ['Single-column stack']),
            ('Breakpoint Rules', ['@media 1024px, 768px, 480px']),
            ('Adaptive Elements', ['Scaled fonts and spacing'])
        ]
    if topic == 'Responsive Design':
        return [
            ('Responsive Header', ['Adaptive nav and branding']),
            ('Fluid Hero', ['Scales with viewport width']),
            ('Adaptive Cards', ['Cards rearrange by screen size']),
            ('Flexible Content', ['No overflow or clipping']),
            ('Responsive Footer', ['Links wrap cleanly'])
        ]
    if topic == 'CSS':
        return [
            ('Theme System', ['Primary/secondary color usage']),
            ('Typography', ['Heading hierarchy + body text']),
            ('Components', ['Buttons, cards, and badges']),
            ('States', ['hover / focus / active styles']),
            ('Tokens', ['Reusable spacing and color variables'])
        ]
    if topic == 'Box Model':
        return [
            ('Margin Zone', ['Outer spacing around component']),
            ('Border Zone', ['Visible component boundary']),
            ('Padding Zone', ['Inner spacing around content']),
            ('Content Zone', ['Text/image content area']),
            ('Sizing Rule', ['box-sizing and width control'])
        ]
    return [
        ('Header', ['Page title and navigation']),
        ('Primary Block', ['Main content for the problem']),
        ('Secondary Block', ['Supporting details']),
        ('Detail Block', ['Additional structured information']),
        ('Footer', ['Final actions or links'])
    ]

for r in range(2, ws.max_row+1):
    level = ws.cell(r,2).value
    topic = ws.cell(r,3).value
    title = ws.cell(r,4).value

    accent,bg = palette(topic)
    rects = layout(f'{level}|{topic}|{title}')
    content = lines_for_problem(topic, title)
    fills = ['#dbeafe','#e2e8f0','#fef3c7','#dcfce7','#ede9fe']

    parts = [
        "<svg xmlns='http://www.w3.org/2000/svg' width='1280' height='720' viewBox='0 0 1280 720'>",
        f"<rect width='1280' height='720' fill='{bg}'/>",
        f"<rect x='24' y='24' width='1232' height='672' rx='20' fill='white' stroke='{accent}' stroke-width='4'/>",
        f"<text x='56' y='78' font-family='Arial, sans-serif' font-size='28' font-weight='800' fill='{accent}'>{title}</text>",
    ]

    for i,(x,y,w,h) in enumerate(rects):
        heading, lines = content[i]
        parts.append(f"<rect x='{x}' y='{y}' width='{w}' height='{h}' rx='12' fill='{fills[i]}'/>")
        parts.append(f"<text x='{x+14}' y='{y+34}' font-family='Arial, sans-serif' font-size='20' font-weight='700' fill='#111827'>{heading}</text>")
        for j,line in enumerate(lines[:4]):
            parts.append(f"<text x='{x+14}' y='{y+58 + j*20}' font-family='Arial, sans-serif' font-size='15' fill='#1f2937'>{line}</text>")

    parts.append('</svg>')

    out_dir = img_root / slugify(level) / slugify(topic)
    out_dir.mkdir(parents=True, exist_ok=True)
    fname = f"q{r-1:03d}-{slugify(title)[:100]}.svg"
    out = out_dir / fname
    out.write_text(''.join(parts), encoding='utf-8')

    rel = out.relative_to(repo).as_posix()
    link = f"https://raw.githubusercontent.com/{owner}/{repo_name}/{branch}/{rel}"
    ws.cell(r,9).value = link
    ws.cell(r,10).value = link

wb.save(xlsx)
print('Updated workbook with strict problem-matched output images:', xlsx)
