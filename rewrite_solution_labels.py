from pathlib import Path
from openpyxl import load_workbook
import re
import hashlib

repo = Path('/Users/user/Frontend-Development-')
wb = load_workbook(repo / 'Frontend_Development_Practice_Bank.xlsx')
solutions_root = repo / 'solutions'

def slugify(text: str) -> str:
    return re.sub(r'[^a-z0-9]+', '-', text.lower()).strip('-')

def palette(topic):
    return {
        'HTML': ('#0f766e','#f0fdfa'),'Semantic HTML': ('#0369a1','#f0f9ff'),'Forms': ('#1d4ed8','#eff6ff'),'Tables': ('#334155','#f8fafc'),
        'CSS': ('#7c3aed','#f5f3ff'),'Box Model': ('#b45309','#fff7ed'),'Flexbox': ('#0e7490','#ecfeff'),'Grid': ('#be185d','#fdf2f8'),
        'Media Queries': ('#166534','#f0fdf4'),'Responsive Design': ('#9a3412','#fff7ed')
    }[topic]

def unique_layout(seed):
    h = int(hashlib.md5(seed.encode()).hexdigest()[:8], 16)
    p = h % 6
    if p == 0:
        return [(80,190,1120,90),(80,295,550,150),(650,295,550,150),(80,460,760,150),(860,460,340,150)]
    if p == 1:
        return [(80,190,1120,90),(80,295,360,315),(460,295,740,95),(460,405,360,205),(840,405,360,205)]
    if p == 2:
        return [(80,190,1120,90),(80,295,1120,120),(80,430,360,180),(460,430,360,180),(840,430,360,180)]
    if p == 3:
        return [(80,190,360,200),(460,190,360,200),(840,190,360,200),(80,410,540,200),(640,410,560,200)]
    if p == 4:
        return [(80,190,760,120),(860,190,340,120),(80,330,360,280),(460,330,360,280),(840,330,360,280)]
    return [(80,190,540,200),(640,190,560,200),(80,410,1120,90),(80,520,550,90),(650,520,550,90)]

def title_core(title: str) -> str:
    # remove suffix after " - topic practice"
    return re.sub(r'\s-\s[^-]*Practice\s[BIH]\d{3}$', '', title).strip()

def labels_for(topic: str, title: str):
    core = title_core(title)
    words = [w for w in re.split(r'[^A-Za-z0-9]+', core) if w]
    # build unique labels directly from question words
    chunks = []
    step = max(1, len(words)//5)
    for i in range(5):
        part = words[i*step:(i+1)*step] if i < 4 else words[i*step:]
        if not part:
            part = words[:1] if words else ['Section']
        chunks.append(' '.join(part[:3]))

    topic_suffix = {
        'HTML': ['Structure','Content','Section','Highlights','Footer'],
        'Semantic HTML': ['Landmark','Article','Aside','Figure','Footer'],
        'Forms': ['Header','Inputs','Choices','Validation','Submit'],
        'Tables': ['Caption','Header','Rows','Summary','Notes'],
        'CSS': ['Theme','Type','State','Component','Tokens'],
        'Box Model': ['Margin','Border','Padding','Content','Spacing'],
        'Flexbox': ['Bar','Row','Actions','Wrap','End'],
        'Grid': ['Top','Left','Center','Right','Bottom'],
        'Media Queries': ['Desktop','Tablet','Mobile','Rule','Adaptive'],
        'Responsive Design': ['Header','Hero','Cards','Flow','Footer'],
    }[topic]

    return [f"{chunks[i]} {topic_suffix[i]}".strip() for i in range(5)]

def make_html(topic, title, seed):
    accent, bg = palette(topic)
    rects = unique_layout(seed)
    labels = labels_for(topic, title)
    fills = ['#dbeafe','#e2e8f0','#fef3c7','#dcfce7','#ede9fe']

    blocks = []
    for i, (x,y,w,h) in enumerate(rects):
        blocks.append(f"<section class='blk' style='left:{x}px;top:{y}px;width:{w}px;height:{h}px;background:{fills[i]};'><h2>{labels[i]}</h2></section>")

    return f"""<!doctype html>
<html lang='en'>
<head>
  <meta charset='utf-8' />
  <meta name='viewport' content='width=device-width, initial-scale=1' />
  <title>{title}</title>
  <style>
    * {{ box-sizing: border-box; }}
    body {{ margin: 0; font-family: Arial, sans-serif; background: {bg}; color: #111827; }}
    .canvas {{ width: 1280px; min-height: 720px; margin: 0 auto; position: relative; }}
    .frame {{ position: absolute; left: 36px; top: 36px; width: 1208px; height: 648px; background: #fff; border: 5px solid {accent}; border-radius: 20px; }}
    .title {{ position: absolute; left: 78px; top: 74px; margin: 0; font-size: 30px; font-weight: 800; color: {accent}; }}
    .sub {{ position: absolute; left: 78px; top: 114px; margin: 0; font-size: 20px; color: #334155; }}
    .blk {{ position: absolute; border-radius: 12px; padding: 10px 14px; overflow: hidden; }}
    .blk h2 {{ margin: 0; font-size: 24px; font-weight: 700; color: #111827; }}
    .foot {{ position: absolute; left: 80px; top: 660px; margin: 0; font-size: 20px; color: #0f172a; }}
    @media (max-width: 1280px) {{
      .canvas {{ width: 100vw; min-height: 56.25vw; }}
      .frame, .title, .sub, .foot, .blk {{ transform-origin: top left; transform: scale(calc(100vw / 1280)); }}
    }}
  </style>
</head>
<body>
  <main class='canvas'>
    <div class='frame'></div>
    <h1 class='title'>{title}</h1>
    <p class='sub'>Expected UI Reference</p>
    {''.join(blocks)}
    <p class='foot'>Recreate this exact final interface.</p>
  </main>
</body>
</html>
"""

for sheet in wb.sheetnames:
    ws = wb[sheet]
    for r in range(2, ws.max_row + 1):
        topic = ws.cell(r, 3).value
        title = ws.cell(r, 4).value
        q = r - 1
        level_slug = slugify(sheet)
        topic_slug = slugify(topic)
        file_slug = slugify(title)[:100]
        path = solutions_root / level_slug / topic_slug / f"q{q:03d}-{file_slug}.html"
        if not path.exists():
            continue
        seed = f"{sheet}|{topic}|{title}"
        path.write_text(make_html(topic, title, seed), encoding='utf-8')

print('Rewrote solution outputs with question-derived unique labels.')
