from pathlib import Path
from openpyxl import load_workbook
import re

repo = Path('/Users/user/Frontend-Development-')
xlsx = Path('/Users/user/Downloads/Frontend_Development_Practice_Bank.xlsx')
wb = load_workbook(xlsx)
ws = wb['Sheet1']

img_root = repo / 'solution-images'
img_root.mkdir(parents=True, exist_ok=True)

owner='Nithya-sri-14'
repo_name='Frontend-Development-'
branch='main'

# columns in current sheet
COL_CLICK = 9
COL_COPY = 10
COL_LEVEL = 2
COL_TOPIC = 3
COL_TITLE = 4

def slugify(t):
    return re.sub(r'[^a-z0-9]+','-',t.lower()).strip('-')

for r in range(2, ws.max_row + 1):
    level = ws.cell(r, COL_LEVEL).value
    topic = ws.cell(r, COL_TOPIC).value
    title = ws.cell(r, COL_TITLE).value

    level_slug = slugify(level)
    topic_slug = slugify(topic)
    file_slug = slugify(title)[:110]

    # create image from same deterministic layout used by solution pages
    # so visual matches expected structure
    import hashlib
    seed = f"{level}|{topic}|{title}"
    h = int(hashlib.md5(seed.encode()).hexdigest()[:8], 16)
    p = h % 6
    if p == 0:
        rects = [(80,190,1120,90),(80,295,550,150),(650,295,550,150),(80,460,760,150),(860,460,340,150)]
    elif p == 1:
        rects = [(80,190,1120,90),(80,295,360,315),(460,295,740,95),(460,405,360,205),(840,405,360,205)]
    elif p == 2:
        rects = [(80,190,1120,90),(80,295,1120,120),(80,430,360,180),(460,430,360,180),(840,430,360,180)]
    elif p == 3:
        rects = [(80,190,360,200),(460,190,360,200),(840,190,360,200),(80,410,540,200),(640,410,560,200)]
    elif p == 4:
        rects = [(80,190,760,120),(860,190,340,120),(80,330,360,280),(460,330,360,280),(840,330,360,280)]
    else:
        rects = [(80,190,540,200),(640,190,560,200),(80,410,1120,90),(80,520,550,90),(650,520,550,90)]

    pal = {
        'HTML': ('#0f766e','#f0fdfa'),'Semantic HTML': ('#0369a1','#f0f9ff'),'Forms': ('#1d4ed8','#eff6ff'),'Tables': ('#334155','#f8fafc'),
        'CSS': ('#7c3aed','#f5f3ff'),'Box Model': ('#b45309','#fff7ed'),'Flexbox': ('#0e7490','#ecfeff'),'Grid': ('#be185d','#fdf2f8'),
        'Media Queries': ('#166534','#f0fdf4'),'Responsive Design': ('#9a3412','#fff7ed')
    }
    accent,bg = pal.get(topic,('#1f2937','#f8fafc'))

    # section titles from question words
    words = [w for w in re.split(r'[^A-Za-z0-9]+', title) if w]
    chunk = max(1, len(words)//5)
    labels=[]
    for i in range(5):
        part = words[i*chunk:(i+1)*chunk] if i<4 else words[i*chunk:]
        if not part:
            part=['Section']
        labels.append(' '.join(part[:3]))

    fills = ['#dbeafe','#e2e8f0','#fef3c7','#dcfce7','#ede9fe']
    body=[]
    for i,(x,y,w,h) in enumerate(rects):
        body.append(f"<rect x='{x}' y='{y}' width='{w}' height='{h}' rx='12' fill='{fills[i]}'/>")
        body.append(f"<text x='{x+14}' y='{y+34}' font-family='Arial, sans-serif' font-size='22' font-weight='700' fill='#111827'>{labels[i]}</text>")

    svg=f"""<svg xmlns='http://www.w3.org/2000/svg' width='1280' height='720' viewBox='0 0 1280 720'>
  <rect width='1280' height='720' fill='{bg}'/>
  <rect x='36' y='36' width='1208' height='648' rx='20' fill='white' stroke='{accent}' stroke-width='5'/>
  <text x='78' y='98' font-family='Arial, sans-serif' font-size='30' font-weight='800' fill='{accent}'>{title}</text>
  {''.join(body)}
</svg>"""

    out_dir = img_root / level_slug / topic_slug
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / f"q{r-1:03d}-{file_slug}.svg"
    out_file.write_text(svg, encoding='utf-8')

    rel = out_file.relative_to(repo).as_posix()
    link = f"https://raw.githubusercontent.com/{owner}/{repo_name}/{branch}/{rel}"
    ws.cell(r, COL_CLICK).value = link
    ws.cell(r, COL_COPY).value = link

wb.save(xlsx)
print('updated workbook with image links:', xlsx)
